from django.db.models.fields import NullBooleanField
from django.shortcuts import render,redirect,get_object_or_404
from django.http import HttpResponse,JsonResponse
from .models import (
    Course,Chapter, GameGoldenFish, Lesson,ExerciseChoiceAnswer,
    TypeExercise,ExerciseArrange,ExerciseWordMissing,ViolympicEndCourse,
    GameWordMemoryCards,ImageOfGame
)
from order.models import OrderItem,Order
from usermember.models import TypeScore,ScoreStudent,ScoreViolympicFinishCourse,StudentCourse
import json 
from django.core import serializers
from django.views import View
from django.views.generic.base import TemplateView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime
from django.utils import timezone
import xlrd
from django.conf import settings
from django.db.models import Avg,Sum,Count
from openpyxl import load_workbook
# Create your views here.


def Deatail_Course(request,course_id):
    cs = Course.objects.get(pk=course_id)
    vio_one_s = ViolympicEndCourse.objects.filter(course_id=course_id)
    std_course_qs = StudentCourse.objects.filter(user_id=request.user.id,course_id=course_id)
    course_learn_qs = StudentCourse.objects.filter(user_id=request.user.id,course_id=course_id)
    if cs:
        chap = Chapter.objects.filter(course_id=cs.id)
        count_chap = chap.count()
        vio_one = None
        std_course = None  
        std_course = None
        learn_course = None  
        if vio_one_s.exists():
            vio_one = vio_one_s[0]
        if std_course_qs.exists():
            std_course = std_course_qs[0]
        if course_learn_qs.exists():
            learn_course = course_learn_qs[0]
        context = {
            'cs':cs,
            'chap':chap,
            'vio_one':vio_one,
            'std_course':std_course,
            'count_chap':count_chap,
            'learn_course':learn_course
        }
        return render(request,'course/course.html',context)   
    else:
        return HttpResponse('Không tìm thay khoa hoc nay')

def checkExistExercise(lesson_id,type_exercise,exercise_answer):
    choice_qs = exercise_answer.objects.filter(lesson_id=lesson_id,type_id=type_exercise)
    if choice_qs:
        choice = choice_qs[0]
        file_excel = choice.file_excel
        if not file_excel:  
            return False
        else:
            return True
    else:
        return False

def checkExistGame(lesson_id,type_game,obj_game):
    qs__ = obj_game.objects.filter(lesson_id=lesson_id,type_id=type_game)
    if qs__:
        return True
    else:
        return False

@login_required(login_url='/login/')
def View_Detail_Lesson(request,course_id,lesson_id):
    less = Lesson.objects.get(pk=lesson_id)  
    # check exist choice 
    choice_ans = checkExistExercise(lesson_id,1,ExerciseChoiceAnswer)
    # check exist word missing 
    missing_ans = checkExistExercise(lesson_id,2,ExerciseWordMissing)
    # check exist arrange word 
    arrange_ans = checkExistExercise(lesson_id,3,ExerciseArrange)
    # check game word memory card 
    game_card = checkExistGame(lesson_id,1,GameWordMemoryCards)
     # check game golden fish
    golden_fish = checkExistGame(lesson_id,2,GameGoldenFish)
    context = {
        'less':less,
        'choice_ans':choice_ans,
        'missing_ans':missing_ans,
        'arrange_ans':arrange_ans,
        'game_card':game_card,
        'golden_fish':golden_fish,
    }
    return render(request,'course/detail_lesson.html',context)

def get_post_data_exercise_choice(request,course_id,lesson_id,num_scores=0,is_doing=0,is_get=0):
    list_choice = ExerciseChoiceAnswer.objects.filter(lesson_id=lesson_id,type_id=1)
    current_user = request.user

    if list_choice:
        choice = list_choice[0]
        file_excel = choice.file_excel
        if not file_excel: 
            
            return JsonResponse({'status':'Filed excel file is null. NOT DATA'})
        else:
            data1 = xlrd.open_workbook(settings.BASE_DIR + choice.file_excel.url) #open a file
            table1 = data1.sheet_by_index(0) #Get a worksheet
            nrows1 = table1.nrows #Rows
            choice.num_rows = nrows1 - 1
            choice.save()

            file_excel = choice.file_excel
            url_file = settings.BASE_DIR + file_excel.url
            data0 = xlrd.open_workbook(url_file) #open a file
            table = data0.sheet_by_index(0) #Get a worksheet
            nrows = table.nrows #Rows
            ncols = table.ncols 
            
            wb = load_workbook(url_file)
            sheet = wb.active
            
            data = []
            for row in sheet.iter_rows(min_row=2, min_col=1, max_row=nrows, max_col=ncols):
                item = {
                    'question':row[0].value,
                    'choice1':row[1].value,
                    'choice2':row[2].value,
                    'choice3':row[3].value,
                    'choice4':row[4].value,
                    'answer':row[5].value
                }
                data.append(item)
            
            getScoreDb = ScoreStudent.objects.filter(user_id=current_user.id,course_id=course_id,lesson_id=lesson_id,type_exercise_id=1,type_score_id=1)
            if(getScoreDb):
                is_doing = 1
                if(is_get == 0): # get du lieu , khong luu
                    num_scores = getScoreDb[0].score
                    
                else:
                    if(is_doing == 1): # co lam bai
                        getScoreDb[0].score = num_scores
                        getScoreDb[0].save()
                        num_scores = getScoreDb[0].score
                        
            else:
                if(is_get == 0): # get du lieu , khong luu
                    is_get = 0
                    is_doing = 0
                    
                else: # post diem
                    if(is_doing == 1):
                        createScoreStd = ScoreStudent(user_id=current_user.id,course_id=course_id,lesson_id=lesson_id,type_exercise_id=1,score=num_scores,type_score_id=1) 
                        createScoreStd.save()    
                
            return JsonResponse({'data':data,'score':num_scores,'is_doing':is_doing,'is_get':is_get})
        
    else:
        return JsonResponse({'status':'Not exist exercise'})

def get_post_data_game_word_card(request,course_id,lesson_id,num_scores=0,is_doing=0,is_get=0):
    list_card = GameWordMemoryCards.objects.filter(lesson_id=lesson_id,type_id=1)
    current_user = request.user
    if list_card:
        card = list_card[0]
        
        
        getScoreDb = ScoreStudent.objects.filter(user_id=current_user.id,course_id=course_id,lesson_id=lesson_id,type_game_id=1,type_score_id=2)
        if(getScoreDb):
            is_doing = 1
            if(is_get == 0): # get du lieu , khong luu
                num_scores = getScoreDb[0].score
                
            else:
                if(is_doing == 1): # co lam bai
                    getScoreDb[0].score = num_scores
                    getScoreDb[0].save()
                    num_scores = getScoreDb[0].score
                    
        else:
            if(is_get == 0): # get du lieu , khong luu
                is_get = 0
                is_doing = 0
                
            else: # post diem
                if(is_doing == 1):
                    createScoreStd = ScoreStudent(user_id=current_user.id,course_id=course_id,lesson_id=lesson_id,type_game_id=1,score=num_scores,type_score_id=2) 
                    createScoreStd.save()
        
        if card:
            data = []
            count = 0
            for value in card.imgs.all():
                item = {
                    'id':count,
                    'img':value.img.url
                }
                data.append(item)
                count +=1
            return JsonResponse({'data' : data,'score':num_scores,'is_doing':is_doing,'is_get':is_get})   
        else:
            return JsonResponse({'status':'Data is null'})
    else:
        return JsonResponse({'status':'Not exist game'})

def get_post_data_exercise_arrange(request,course_id,lesson_id,num_scores=0,is_doing=0,is_get=0):
    list_ar = ExerciseArrange.objects.filter(lesson_id=lesson_id).filter(type_id=3)
    current_user = request.user
    
    if list_ar:
        arr = list_ar[0]
        file_excel = arr.file_excel
        if not file_excel: 
            
            return JsonResponse({'status':'Filed excel file is null. NOT DATA'})
        else:
            data1 = xlrd.open_workbook(settings.BASE_DIR + arr.file_excel.url) #open a file
            table1 = data1.sheet_by_index(0) #Get a worksheet
            nrows1 = table1.nrows #Rows
            arr.num_rows = nrows1 - 1
            arr.save()


            file_excel = arr.file_excel
            url_file =settings.BASE_DIR + file_excel.url
            data0= xlrd.open_workbook(url_file) #open a file
            table = data0.sheet_by_index(0) #Get a worksheet
            nrows = table.nrows #Rows
            ncols = table.ncols 
            
            wb = load_workbook(url_file)
            sheet = wb.active
            
            data = []
            for row in sheet.iter_rows(min_row=2, min_col=1, max_row=nrows, max_col=ncols):
                item = {
                    'question':row[0].value
                }
                data.append(item)
            
            getScoreDb = ScoreStudent.objects.filter(user_id=current_user.id,course_id=course_id,lesson_id=lesson_id,type_exercise_id=3,type_score_id=1)
            if(getScoreDb):
                is_doing = 1
                if(is_get == 0): # get du lieu , khong luu
                    num_scores = getScoreDb[0].score
                    
                else:
                    if(is_doing == 1): # co lam bai
                        getScoreDb[0].score = num_scores
                        getScoreDb[0].save()
                        num_scores = getScoreDb[0].score
            else:
                if(is_get == 0): # get du lieu , khong luu
                    is_get = 0
                    is_doing = 0
                else: # post diem
                    if(is_doing == 1):
                        createScoreStd = ScoreStudent(user_id=current_user.id,course_id=course_id,lesson_id=lesson_id,type_exercise_id=3,score=num_scores,type_score_id=1) 
                        createScoreStd.save()
            
            return JsonResponse({'data' : data,'score':num_scores,'is_doing':is_doing,'is_get':is_get})
        
    else:
        return JsonResponse({'status':'Not exist exercise'})


def get_post_data_exercise_missing(request,course_id,lesson_id,num_scores=0,is_doing=0,is_get=0):
    list_miss = ExerciseWordMissing.objects.filter(lesson_id=lesson_id).filter(type_id=2)
    current_user = request.user
    
    if list_miss:
        miss = list_miss[0]
        file_excel = miss.file_excel
        if not file_excel:
            
            return JsonResponse({'status':'Filed excel file is null. NOT DATA'})
        else:
            data1 = xlrd.open_workbook(settings.BASE_DIR + miss.file_excel.url) #open a file
            table1 = data1.sheet_by_index(0) #Get a worksheet
            nrows1 = table1.nrows #Rows
            miss.num_rows = nrows1 - 1
            miss.save()  

            file_excel = miss.file_excel

            url_file =settings.BASE_DIR + file_excel.url
            data0 = xlrd.open_workbook(url_file) #open a file
            table = data0.sheet_by_index(0) #Get a worksheet
            nrows = table.nrows #Rows
            ncols = table.ncols 
            
            wb = load_workbook(url_file)
            sheet = wb.active
            
            data = []
            for row in sheet.iter_rows(min_row=2, min_col=1, max_row=nrows, max_col=ncols):
                item = {
                    'question':row[0].value,
                    'answer':row[1].value
                }
                data.append(item)
            
            getScoreDb = ScoreStudent.objects.filter(user_id=current_user.id,course_id=course_id,lesson_id=lesson_id,type_exercise_id=2,type_score_id=1)
            if(getScoreDb):
                is_doing = 1
                if(is_get == 0): # get du lieu , khong luu
                    num_scores = getScoreDb[0].score
                    
                else:
                    if(is_doing == 1): # co lam bai
                        getScoreDb[0].score = num_scores
                        getScoreDb[0].save()
                        num_scores = getScoreDb[0].score
                        
            else:
                if(is_get == 0): # get du lieu , khong luu
                    is_get = 0
                    is_doing = 0
                    
                else: # post diem
                    if(is_doing == 1):
                        createScoreStd = ScoreStudent(user_id=current_user.id,course_id=course_id,lesson_id=lesson_id,type_exercise_id=2,score=num_scores,type_score_id=1) 
                        createScoreStd.save()
                        
            
           
            return JsonResponse({'data' : data,'score':num_scores,'is_doing':is_doing,'is_get':is_get})
      
    else:
        return JsonResponse({'status':'Not exist exercise'})
    
 
@login_required(login_url='/login/')
@csrf_exempt
def Violympic_End_Course(request,course_id,vio_one_id):
    vio = ViolympicEndCourse.objects.filter(pk=vio_one_id).filter(course_id=course_id)
    history = ScoreViolympicFinishCourse.objects.filter(user_id=request.user.id,course_id=course_id).order_by('-id')
    if request.method == 'POST':
        score = request.POST.get('score',None) 
        time_start = request.POST.get('time_start',None) 
        time_finish= request.POST.get('time_finish',None)

        c_time_start = datetime.strptime(time_start, "%Y-%m-%d %H:%M:%S")
        
        current_tz = timezone.get_current_timezone()
        tz_time_start = current_tz.localize(c_time_start)

        c_time_finish = datetime.strptime(time_finish, "%Y-%m-%d %H:%M:%S")
        
        tz_time_finish = current_tz.localize(c_time_finish)


        instance  =  ScoreViolympicFinishCourse(
            user_id=request.user.id,
            course_id=course_id,
            score=score,
            time_start=tz_time_start,
            time_finish=tz_time_finish
            )

        instance.save()
        sco_vio = ScoreViolympicFinishCourse.objects.values().get(pk=instance.id)
        historyes = ScoreViolympicFinishCourse.objects.values().filter(user_id=request.user.id,course_id=course_id).order_by('-id')
        list_history = list(historyes)
        return JsonResponse({'status':'Your Post Success Data','data_post':sco_vio,'list_history':list_history})

    
    if vio:
        return render(request,'course/violympic_end_course.html',{'vio':vio[0],'history':history})
    else:
         return redirect('home:index')


def get_post_data_game_golden_fish(request,course_id,lesson_id,num_scores=0,is_doing=0,is_get=0):
    list_qs = GameGoldenFish.objects.filter(lesson_id=lesson_id,type_id=2)
    current_user = request.user
    
    if list_qs:
        game = list_qs[0]
        file_excel = game.file_excel
        if not file_excel:
            
            return JsonResponse({'status':'Filed excel file is null. NOT DATA'})
        else:
            data1 = xlrd.open_workbook(settings.BASE_DIR + game.file_excel.url) #open a file
            table1 = data1.sheet_by_index(0) #Get a worksheet
            nrows1 = table1.nrows #Rows
            game.num_rows = nrows1 - 1
            game.save()  

            file_excel = game.file_excel

            url_file =settings.BASE_DIR + file_excel.url
            data0 = xlrd.open_workbook(url_file) #open a file
            table = data0.sheet_by_index(0) #Get a worksheet
            nrows = table.nrows #Rows
            ncols = table.ncols 
            
            wb = load_workbook(url_file)
            sheet = wb.active
            
            data = []
            for row in sheet.iter_rows(min_row=2, min_col=1, max_row=nrows, max_col=ncols):
                item = {
                    'question':row[0].value,
                    'answer1':row[1].value,
                    'answer2':row[2].value,
                    'answer3':row[3].value,
                }
                data.append(item)
            
            getScoreDb = ScoreStudent.objects.filter(user_id=current_user.id,course_id=course_id,lesson_id=lesson_id,type_game_id=2,type_score_id=2)
            if(getScoreDb):
                is_doing = 1
                if(is_get == 0): # get du lieu , khong luu
                    num_scores = getScoreDb[0].score
                    
                else:
                    if(is_doing == 1): # co lam bai
                        getScoreDb[0].score = num_scores
                        getScoreDb[0].save()
                        num_scores = getScoreDb[0].score
                        
            else:
                if(is_get == 0): # get du lieu , khong luu
                    is_get = 0
                    is_doing = 0
                    
                else: # post diem
                    if(is_doing == 1):
                        createScoreStd = ScoreStudent(user_id=current_user.id,course_id=course_id,lesson_id=lesson_id,type_game_id=2,score=num_scores,type_score_id=2) 
                        createScoreStd.save()          
           
            return JsonResponse({'data' : data,'score':num_scores,'is_doing':is_doing,'is_get':is_get})
      
    else:
        return JsonResponse({'status':'Not exist exercise'})
    





